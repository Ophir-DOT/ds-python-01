from __future__ import annotations

from openpyxl import load_workbook

from ds_tool.excel.export import export_workbook
from ds_tool.models import (
    FieldPermission,
    FieldSpec,
    ObjectGeneralInfo,
    ObjectSpec,
    ProfileSpec,
)


def test_export_creates_expected_sheets(tmp_path) -> None:
    spec = ObjectSpec(
        general=ObjectGeneralInfo(api_name="Account", label="Account", plural_label="Accounts"),
        fields=[FieldSpec(api_name="Name", label="Name", type="string", required=True)],
        profiles=[
            ProfileSpec(
                full_name="Admin",
                label="System Administrator",
                kind="Profile",
                field_permissions=[FieldPermission(field="Account.Name", readable=True, editable=True)],
            )
        ],
    )
    out = export_workbook([spec], tmp_path / "spec.xlsx")
    assert out.exists()

    wb = load_workbook(out)
    assert "Overview" in wb.sheetnames
    assert "Account Fields" in wb.sheetnames
    assert "Account FLS" in wb.sheetnames
    # header row written
    assert wb["Account Fields"]["A1"].value == "Field"
    assert wb["Account Fields"]["A2"].value == "Name"
