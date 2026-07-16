"""Excel export of ObjectSpec(s) — one workbook, a sheet per section.

Mirrors the Aura "Excel Export" action. Reuses the ObjectSpec models directly.
"""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models import ObjectSpec

_HEADER_FONT = Font(bold=True)


def _sheet(wb: Workbook, title: str) -> Worksheet:
    # Excel caps sheet titles at 31 chars and forbids []:*?/\.
    safe = title[:31]
    for ch in r"[]:*?/\\":
        safe = safe.replace(ch, "-")
    return wb.create_sheet(safe)


def _write(ws: Worksheet, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for row in rows:
        ws.append(["" if v is None else v for v in row])
    for col_idx, header in enumerate(headers, start=1):
        width = max(len(str(header)), 12)
        for row in rows:
            if col_idx <= len(row) and row[col_idx - 1] is not None:
                width = max(width, min(len(str(row[col_idx - 1])), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def _yn(value: bool) -> str:
    return "Yes" if value else ""


def export_workbook(specs: Sequence[ObjectSpec], path: Path) -> Path:
    """Write one .xlsx covering every spec. Section sheets are prefixed per object."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    overview = wb.create_sheet("Overview")
    _write(
        overview,
        ["API Name", "Label", "Plural", "Sharing Model", "History Tracking"],
        [
            [
                s.general.api_name,
                s.general.label,
                s.general.plural_label,
                s.general.sharing_model,
                _yn(s.general.history_tracking_enabled),
            ]
            for s in specs
        ],
    )

    for spec in specs:
        obj = spec.general.api_name

        _write(
            _sheet(wb, f"{obj} Fields"),
            ["Field", "API Name", "Type", "Required", "Unique", "Formula", "Help Text"],
            [
                [f.label, f.api_name, f.type, _yn(f.required), _yn(f.unique), f.formula, f.help_text]
                for f in spec.fields
            ],
        )

        fls_rows: list[list] = []
        for p in spec.profiles:
            for fp in p.field_permissions:
                fls_rows.append([p.label, p.kind, fp.field, _yn(fp.readable), _yn(fp.editable)])
        _write(
            _sheet(wb, f"{obj} FLS"),
            ["Profile/Perm Set", "Kind", "Field", "Read", "Edit"],
            fls_rows,
        )

        obj_perm_rows: list[list] = []
        for p in spec.profiles:
            for op in p.object_permissions:
                obj_perm_rows.append(
                    [
                        p.label,
                        p.kind,
                        _yn(op.read),
                        _yn(op.create),
                        _yn(op.edit),
                        _yn(op.delete),
                        _yn(op.view_all),
                        _yn(op.modify_all),
                    ]
                )
        _write(
            _sheet(wb, f"{obj} Object Perms"),
            ["Profile/Perm Set", "Kind", "Read", "Create", "Edit", "Delete", "View All", "Modify All"],
            obj_perm_rows,
        )

        _write(
            _sheet(wb, f"{obj} Record Types"),
            ["API Name", "Label", "Active", "Description"],
            [[rt.api_name, rt.label, _yn(rt.active), rt.description] for rt in spec.record_types],
        )

        _write(
            _sheet(wb, f"{obj} Validation Rules"),
            ["Rule", "Active", "Error Condition", "Error Message", "Description"],
            [
                [v.api_name, _yn(v.active), v.error_condition, v.error_message, v.description]
                for v in spec.validation_rules
            ],
        )

        _write(
            _sheet(wb, f"{obj} Triggers"),
            ["Name", "Status", "API Version", "Events"],
            [[t.name, t.status, t.api_version, ", ".join(t.events)] for t in spec.apex_triggers],
        )

        _write(
            _sheet(wb, f"{obj} Flows"),
            ["Label", "API Name", "Process Type", "Status", "Description"],
            [[f.label, f.api_name, f.process_type, f.status, f.description] for f in spec.flows],
        )

        _write(
            _sheet(wb, f"{obj} Email Alerts"),
            ["Name", "Description", "Template", "Sender Type", "Recipients"],
            [
                [a.api_name, a.description, a.template, a.sender_type, ", ".join(a.recipients)]
                for a in spec.email_alerts
            ],
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path
